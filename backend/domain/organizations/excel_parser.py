from __future__ import annotations

import io
from typing import Any, BinaryIO
import openpyxl

from .validators import clean_text


def parse_organizations_excel(file_obj: BinaryIO) -> list[dict[str, Any]]:
    # Open the workbook using openpyxl from the memory buffer / file stream
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    if not wb.sheetnames:
        raise ValueError("Excel file is empty.")
    
    # We parse the first sheet containing SIB enterprise data
    sheet = wb.worksheets[0]
    
    # Ensure there is at least header rows and one data row
    if sheet.max_row < 3:
        raise ValueError("Excel file must contain at least 2 header rows and at least 1 data row.")
        
    records: list[dict[str, Any]] = []
    for row in range(3, sheet.max_row + 1):
        trade_name = sheet.cell(row=row, column=1).value
        registered_name = sheet.cell(row=row, column=2).value
        
        # If both trade name and registered name are empty, treat as an empty row and skip
        cleaned_trade_name = clean_text(trade_name)
        cleaned_registered_name = clean_text(registered_name)
        if cleaned_trade_name is None and cleaned_registered_name is None:
            continue
        
        def get_str(col: int) -> str | None:
            val = sheet.cell(row=row, column=col).value
            return clean_text(val)

        def get_year(col: int) -> int | None:
            val = sheet.cell(row=row, column=col).value
            if val is None:
                return None
            try:
                return int(float(val))
            except (ValueError, TypeError):
                return None

        # contacts
        website = get_str(7)
        email = get_str(8)
        phone = get_str(9)
        
        # has positive social impact (Col 21)
        social_impact_raw = get_str(21)
        has_social_impact: bool | None = None
        if social_impact_raw:
            sil = social_impact_raw.lower()
            if "có" in sil or "co" in sil or "1" in sil or "true" in sil:
                has_social_impact = True
            elif "không" in sil or "khong" in sil or "2" in sil or "false" in sil:
                has_social_impact = False

        # environmental impact areas (Col 17 to 20)
        env_areas: list[str | None] = []
        for col in [17, 18, 19, 20]:
            val = get_str(col)
            if val:
                env_areas.append(val)
        while len(env_areas) < 4:
            env_areas.append(None)

        # other industry sectors (Col 14 to 16)
        other_sectors: list[str | None] = []
        for col in [14, 15, 16]:
            val = get_str(col)
            other_sectors.append(val)

        record = {
            "id": f"ROW_{row}",
            "general": {
                "tradeName": cleaned_trade_name,
                "registeredName": cleaned_registered_name,
                "foundedYear": get_year(3),
                "taxCode": get_str(4),
                "location": {
                    "province": get_str(5),
                    "ward": get_str(6)
                },
                "contacts": {
                    "website": website,
                    "email": email,
                    "phone": phone
                },
                "operationalStatus": get_str(10),
                "closedYear": get_year(11)
            },
            "classification": {
                "organizationType": get_str(12),
                "primaryIndustrySector": get_str(13),
                "otherIndustrySectors": other_sectors,
                "environmentalImpactAreas": env_areas,
                "hasPositiveSocialImpact": has_social_impact,
                "primaryProductType": get_str(22),
                "otherProductType": get_str(23)
            }
        }
        records.append(record)
        
    return records
